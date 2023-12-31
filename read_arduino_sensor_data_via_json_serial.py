import serial
import json
from openpyxl import Workbook, load_workbook
from datetime import datetime
import time

# Define the COM port for Arduino (e.g., COM9)
COM_PORT = 'COM5'

# Define the path to your Excel file
excel_file_path = r'C:\Users\bushb\Documents\CO2-RH-TEMPC.xlsx'

# Create or load the Excel workbook
try:
    workbook = load_workbook(excel_file_path)
except FileNotFoundError:  # If no file exists, create one
    workbook = Workbook()
    workbook.save(excel_file_path)
    print(f"Created a new Excel file: {excel_file_path}")
except Exception as e:
    print(f"Error opening Excel file: {str(e)}")
    exit(1)

# Select the default sheet (you can change the sheet name as needed)
sheet = workbook.active

# Initialize the serial connection
ser = None

while True:
    # Attempt to establish a serial connection
    try:
        if ser is None or not ser.is_open:
            ser = serial.Serial(COM_PORT, 115200, timeout=1)
            print(f"Connected to {COM_PORT}")
    except Exception as e:
        print(f"Failed to connect to {COM_PORT}: {str(e)}")
        time.sleep(1)  # Wait for 1 second before retrying
        continue  # Continue to the next iteration of the loop

    # Main loop for reading data from Arduino
    while True:
        try:
            # Read a line from the Arduino
            line = ser.readline().decode('utf-8').strip()

            sensor_data = json.loads(line)

            # Extract data from the JSON object
            id = sensor_data["ID"]
            co2 = sensor_data["co2"]
            humidity = sensor_data["%RH"]
            rhsp = sensor_data["RHSP"]
            box_temperature = sensor_data["boxTempC"]
            bhsp = sensor_data["BHSP"]
            water_temperature = sensor_data["waterTempC"]
            ihsp = sensor_data["IHSP"]

            # Get the current date and time
            current_time = datetime.now().strftime('%Y-%m-%d %H:%M:%S')

            # Append data to the Excel sheet
            sheet.append([current_time, co2, humidity, rhsp, box_temperature, bhsp, water_temperature, ihsp])

            # Save the Excel file
            workbook.save(excel_file_path)

            print(f"Data recorded: {current_time}, CO2: {co2}, RH: {humidity}, RHSP: {rhsp}, Box Temperature: {box_temperature}, BHSP: {bhsp}, Water Temperature: {water_temperature}, IHSP: {ihsp}")
        except json.JSONDecodeError:
            print(f"Invalid JSON data: {line}")
        except KeyError as e:
            print(f"Missing key in JSON data: {e}")
        except Exception as e:
            print(f"Error processing data: {str(e)}")
            # Continue to the next iteration of the loop on error
            continue

    # Close the serial connection when the inner loop is exited
    ser.close()
    print("Serial connection closed. Retrying in 1 second.")
    time.sleep(1)  # Wait for 1 second before attempting to reconnect
